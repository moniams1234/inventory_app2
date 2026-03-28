"""
export.py – Eksport wyników do pliku Excel z formatowaniem biznesowym.
"""
from __future__ import annotations

import io
from datetime import date
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# Paleta kolorów szaro-pomarańczowa
ORANGE = "E8650A"
LIGHT_ORANGE = "FAD7B8"
DARK_GREY = "404040"
MID_GREY = "808080"
LIGHT_GREY = "F2F2F2"
WHITE = "FFFFFF"


def _header_style(ws, row_num: int, fill_color: str = ORANGE) -> None:
    """Nadaje styl nagłówkowy całemu wierszowi."""
    for cell in ws[row_num]:
        cell.fill = PatternFill("solid", fgColor=fill_color)
        cell.font = Font(bold=True, color=WHITE, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            bottom=Side(style="medium", color=DARK_GREY),
            right=Side(style="thin", color=MID_GREY),
        )


def _set_column_widths(ws, widths: dict[int, int]) -> None:
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _alt_row_fill(ws, start_row: int, end_row: int, max_col: int) -> None:
    """Naprzemienne tło wierszy dla czytelności."""
    for r in range(start_row, end_row + 1):
        fill = PatternFill("solid", fgColor=LIGHT_GREY if r % 2 == 0 else WHITE)
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            if cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF", WHITE, "FF" + WHITE):
                cell.fill = fill


def export_to_excel(
    df: pd.DataFrame,
    summary: pd.DataFrame,
    analysis_date: date,
    stats: dict[str, Any],
    warnings_list: list[str],
    errors_list: list[str],
    mapping_source_label: str,
) -> bytes:
    """
    Generuje plik Excel z trzema arkuszami:
      1. Dane szczegółowe
      2. Podsumowanie
      3. Log walidacji

    Zwraca bajty gotowe do pobrania.
    """
    output = io.BytesIO()

    # ------------------------------------------------------------------ #
    # Arkusz 1 – Dane szczegółowe                                         #
    # ------------------------------------------------------------------ #
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        wb = writer.book

        # Formaty XlsxWriter
        fmt_header = wb.add_format({
            "bold": True,
            "font_color": WHITE,
            "bg_color": "#" + ORANGE,
            "border": 1,
            "align": "center",
            "valign": "vcenter",
            "text_wrap": True,
        })
        fmt_pct = wb.add_format({"num_format": "0%", "border": 1})
        fmt_num = wb.add_format({"num_format": "#,##0.00", "border": 1})
        fmt_text = wb.add_format({"border": 1})
        fmt_alt = wb.add_format({"bg_color": "#" + LIGHT_GREY, "border": 1})
        fmt_pct_alt = wb.add_format({"num_format": "0%", "bg_color": "#" + LIGHT_GREY, "border": 1})
        fmt_num_alt = wb.add_format({"num_format": "#,##0.00", "bg_color": "#" + LIGHT_GREY, "border": 1})
        fmt_total = wb.add_format({
            "bold": True,
            "bg_color": "#" + LIGHT_ORANGE,
            "num_format": "#,##0.00",
            "border": 1,
        })
        fmt_total_pct = wb.add_format({
            "bold": True,
            "bg_color": "#" + LIGHT_ORANGE,
            "num_format": "0%",
            "border": 1,
        })

        # ---- Dane szczegółowe ----
        df_export = df.copy()
        df_export.to_excel(writer, sheet_name="Dane szczegółowe", index=False, startrow=1)
        ws_det = writer.sheets["Dane szczegółowe"]

        # Tytuł
        title_fmt = wb.add_format({
            "bold": True, "font_size": 13, "font_color": "#" + ORANGE,
            "valign": "vcenter",
        })
        ws_det.write(0, 0, f"Wiekowanie zapasów – dane szczegółowe | Data analizy: {analysis_date.strftime('%d.%m.%Y')} | Mapping: {mapping_source_label}", title_fmt)

        ncols = len(df_export.columns)
        # Nagłówki
        for col_i, col_name in enumerate(df_export.columns):
            ws_det.write(1, col_i, col_name, fmt_header)
        ws_det.set_row(1, 30)

        # Dane
        pct_cols = {"% rezerwy"}
        num_cols = {"Wartość mag.", "Kwota rezerwy", "Kurs DKK", "Wartość DKK", "Stan mag."}

        for row_i, row_data in enumerate(df_export.itertuples(index=False), start=2):
            is_alt = row_i % 2 == 0
            for col_i, col_name in enumerate(df_export.columns):
                val = row_data[col_i]
                if col_name in pct_cols:
                    fmt = fmt_pct_alt if is_alt else fmt_pct
                elif col_name in num_cols:
                    fmt = fmt_num_alt if is_alt else fmt_num
                else:
                    fmt = fmt_alt if is_alt else fmt_text
                # Konwertuj NaT/NaN na pusty string
                if val is pd.NaT or (isinstance(val, float) and pd.isna(val)):
                    val = ""
                ws_det.write(row_i, col_i, val, fmt)

        # Szerokości kolumn
        col_widths = {
            "Index materiałowy": 18,
            "Magazyn": 22,
            "Typ surowca": 22,
            "Data przyjęcia": 14,
            "Wartość mag.": 14,
            "Rodzaj indeksu": 15,
            "Type of materials": 16,
            "Przedział wiekowania": 16,
            "% rezerwy": 11,
            "Status pozycji": 13,
            "Kwota rezerwy": 14,
        }
        for col_i, col_name in enumerate(df_export.columns):
            width = col_widths.get(col_name, max(len(str(col_name)) + 2, 10))
            ws_det.set_column(col_i, col_i, width)

        ws_det.freeze_panes(2, 0)
        ws_det.autofilter(1, 0, 1 + len(df_export), ncols - 1)

        # ---- Podsumowanie ----
        ws_sum = writer.book.add_worksheet("Podsumowanie")

        title_row = 0
        ws_sum.write(title_row, 0, f"Tabela podsumowująca rezerw | Data analizy: {analysis_date.strftime('%d.%m.%Y')}", title_fmt)

        # Wypłaszcz nagłówki MultiIndex
        summary_flat = summary.copy()
        if isinstance(summary_flat.columns, pd.MultiIndex):
            summary_flat.columns = [" | ".join(str(c) for c in col).strip(" | ")
                                     for col in summary_flat.columns]
        summary_flat = summary_flat.reset_index()

        header_row = 2
        for c_idx, col_name in enumerate(summary_flat.columns):
            ws_sum.write(header_row, c_idx, col_name, fmt_header)
        ws_sum.set_row(header_row, 40)

        for r_idx, row_data in enumerate(summary_flat.itertuples(index=False), start=header_row + 1):
            is_total = str(row_data[0]) == "SUMA KOŃCOWA"
            for c_idx, val in enumerate(row_data):
                if is_total:
                    f = fmt_total_pct if "%" in summary_flat.columns[c_idx] else fmt_total
                elif isinstance(val, (int, float)) and not isinstance(val, bool):
                    f = fmt_num_alt if r_idx % 2 == 0 else fmt_num
                else:
                    f = fmt_alt if r_idx % 2 == 0 else fmt_text
                if val is pd.NaT or (isinstance(val, float) and pd.isna(val)):
                    val = 0
                ws_sum.write(r_idx, c_idx, val, f)

        # Szerokości
        ws_sum.set_column(0, 0, 28)
        for c_idx in range(1, len(summary_flat.columns)):
            ws_sum.set_column(c_idx, c_idx, 20)

        ws_sum.freeze_panes(header_row + 1, 1)
        ws_sum.autofilter(header_row, 0, header_row + len(summary_flat), len(summary_flat.columns) - 1)

        # Metryki ogólne w arkuszu Podsumowanie (na dole)
        metrics_start = header_row + len(summary_flat) + 4
        metrics_header_fmt = wb.add_format({
            "bold": True, "font_color": WHITE, "bg_color": "#" + DARK_GREY,
            "border": 1, "align": "center",
        })
        metrics_val_fmt = wb.add_format({"num_format": "#,##0.00", "border": 1, "bg_color": "#" + LIGHT_GREY})

        ws_sum.write(metrics_start, 0, "Metryki ogólne", metrics_header_fmt)
        ws_sum.write(metrics_start, 1, "Wartość", metrics_header_fmt)
        metrics = [
            ("Łączna liczba rekordów", stats.get("total", 0)),
            ("Zmapowane (Mapp2)", stats.get("mapped", 0)),
            ("Niezmapowane (UNMAPPED)", stats.get("unmapped", 0)),
            ("Błędne daty", stats.get("date_errors", 0)),
            ("Rekordy z rezerwą > 0", stats.get("with_reserve", 0)),
            ("Łączna kwota rezerwy (PLN)", stats.get("total_reserve", 0)),
            ("Łączna wartość magazynowa (PLN)", stats.get("total_value", 0)),
        ]
        for i, (label, value) in enumerate(metrics, start=metrics_start + 1):
            ws_sum.write(i, 0, label, fmt_text)
            ws_sum.write(i, 1, value, metrics_val_fmt)

        # ---- Log walidacji ----
        ws_log = writer.book.add_worksheet("Log walidacji")
        ws_log.write(0, 0, "Log walidacji i ostrzeżeń", title_fmt)
        ws_log.write(1, 0, "Typ", fmt_header)
        ws_log.write(1, 1, "Opis", fmt_header)
        ws_log.set_column(0, 0, 15)
        ws_log.set_column(1, 1, 80)

        log_row = 2
        err_fmt = wb.add_format({"font_color": "CC0000", "border": 1})
        warn_fmt = wb.add_format({"font_color": "CC6600", "border": 1})
        ok_fmt = wb.add_format({"font_color": "006600", "border": 1})

        for e in errors_list:
            ws_log.write(log_row, 0, "BŁĄD", err_fmt)
            ws_log.write(log_row, 1, e, err_fmt)
            log_row += 1
        for w in warnings_list:
            ws_log.write(log_row, 0, "OSTRZEŻENIE", warn_fmt)
            ws_log.write(log_row, 1, w, warn_fmt)
            log_row += 1
        if log_row == 2:
            ws_log.write(log_row, 0, "OK", ok_fmt)
            ws_log.write(log_row, 1, "Brak błędów i ostrzeżeń.", ok_fmt)

    output.seek(0)
    return output.read()


def df_to_csv_bytes(df: pd.DataFrame) -> bytes:
    """Konwertuje DataFrame do CSV jako bajty UTF-8 z BOM (dla Excel)."""
    return df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")


def summary_to_csv_bytes(summary: pd.DataFrame) -> bytes:
    """Konwertuje tabelę podsumowującą do CSV."""
    flat = summary.copy()
    if isinstance(flat.columns, pd.MultiIndex):
        flat.columns = [" | ".join(str(c) for c in col).strip(" | ")
                        for col in flat.columns]
    flat = flat.reset_index()
    return flat.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")
